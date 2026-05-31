"""
Job Card Excel Exporter - Aggregated by Request Number
Exports jobs grouped and aggregated by request number with pricing calculations
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class JobCardExporter:
    """Handles Excel export for job cards with aggregation by request number"""
    
    # Pricing constants
    FIXED_RATE = 200  # For jobs with PCS < 5
    RATE_PER_PCS = 45  # For jobs with PCS >= 5
    GST_RATE = 0.18  # 18% GST
    CGST_RATE = 0.09  # 9% CGST
    SGST_RATE = 0.09  # 9% SGST
    
    def __init__(self, output_dir="exports"):
        """Initialize exporter with output directory"""
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    @staticmethod
    def calculate_rate(pcs):
        """Calculate rate based on piece count
        
        Args:
            pcs (int): Number of pieces
            
        Returns:
            int: Rate per piece or fixed rate
        """
        if pcs < 5:
            return JobCardExporter.FIXED_RATE
        else:
            return JobCardExporter.RATE_PER_PCS
    
    @staticmethod
    def calculate_amount(pcs):
        """Calculate base amount
        
        Args:
            pcs (int): Number of pieces
            
        Returns:
            float: Amount for this job
        """
        rate = JobCardExporter.calculate_rate(pcs)
        if pcs < 5:
            return rate  # Fixed 200
        else:
            return rate * pcs  # 45 × pcs
    
    @staticmethod
    def setup_styles():
        """Create and return style definitions"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=10)
        
        gst_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        gst_font = Font(bold=True, size=10)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        return {
            "header_fill": header_fill,
            "header_font": header_font,
            "subheader_fill": subheader_fill,
            "subheader_font": subheader_font,
            "total_fill": total_fill,
            "total_font": total_font,
            "gst_fill": gst_fill,
            "gst_font": gst_font,
            "border": border,
            "center_alignment": center_alignment,
        }
    
    def export_to_excel(self, jobs_data, request_details):
        """Export jobs with per-job details
        
        Args:
            jobs_data (list): List of job dicts with keys:
                - job_id, request_no, date, licence, jeweller, pcs, weight, 
                  scrap_weight, current_weight, purity, bill_no
            request_details (dict): Dict with keys (for compatibility, not used for per-job data)
                - license_no, date (optional)
                
        Returns:
            str: Path to created Excel file
        """
        if not jobs_data:
            raise ValueError("No jobs data provided")
        
        # Use jobs directly without aggregation - keep individual job details
        jobs_list = []
        for job in jobs_data:
            jobs_list.append({
                "job_id": job.get("job_id", "N/A"),
                "request_no": job.get("request_no", "N/A"),
                "date": job.get("date", datetime.now().strftime("%d-%m-%Y")),
                "licence": job.get("licence", "N/A"),
                "jeweller": job.get("jeweller", "N/A"),
                "pcs": int(job.get("pcs", 0)),
                "weight": float(job.get("weight", 0)),
                "scrap_weight": float(job.get("scrap_weight", 0)),
                "current_weight": float(job.get("current_weight", 0)),
                "purity": job.get("purity", ""),
                "bill_no": job.get("bill_no", "")
            })
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Details"
        
        # Get styles
        styles = self.setup_styles()
        
        # Set column widths
        column_widths = [12, 15, 15, 15, 12, 18, 10, 12, 15, 18, 12, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Header section
        row = 1
        ws[f"A{row}"] = "JOB EXPORT REPORT"
        ws[f"A{row}"].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f"A{row}"].fill = styles["header_fill"]
        ws.merge_cells(f"A{row}:M{row}")
        ws[f"A{row}"].alignment = styles["center_alignment"]
        
        # Column headers
        row = 3
        headers = ["Date", "Bill No", "License No", "Request No", "Job ID", "Jeweller", "PCS", 
                   "Weight(g)", "Scrap Weight(g)", "Current Weight(g)", "Purity", "Base Amount", 
                   "GST (18%)", "CGST (9%)", "SGST (9%)", "Total Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = styles["subheader_fill"]
            cell.font = styles["subheader_font"]
            cell.alignment = styles["center_alignment"]
            cell.border = styles["border"]
        
        # Request data rows - ONE ROW PER JOB with detail
        row = 4
        grand_total_pcs = 0
        grand_total_weight = 0
        grand_total_scrap = 0
        grand_total_current = 0
        grand_total_base_amount = 0
        grand_total_gst = 0
        grand_total_cgst = 0
        grand_total_sgst = 0
        
        for job in jobs_list:
            pcs = job["pcs"]
            weight = job["weight"]
            scrap_weight = job["scrap_weight"]
            current_weight = job["current_weight"]
            
            # Calculate base amount on PCS
            base_amount = self.calculate_amount(pcs)
            gst_amount = base_amount * self.GST_RATE
            cgst_amount = base_amount * self.CGST_RATE
            sgst_amount = base_amount * self.SGST_RATE
            total_amount = base_amount + gst_amount
            
            # Round all amounts to 2 decimals
            base_amount = round(base_amount, 2)
            gst_amount = round(gst_amount, 2)
            cgst_amount = round(cgst_amount, 2)
            sgst_amount = round(sgst_amount, 2)
            total_amount = round(total_amount, 2)
            
            # Update grand totals
            grand_total_pcs += pcs
            grand_total_weight += weight
            grand_total_scrap += scrap_weight
            grand_total_current += current_weight
            grand_total_base_amount += base_amount
            grand_total_gst += gst_amount
            grand_total_cgst += cgst_amount
            grand_total_sgst += sgst_amount
            
            # Write job data - ONE ROW PER JOB
            ws.cell(row=row, column=1, value=job["date"])
            ws.cell(row=row, column=2, value=job["bill_no"])
            ws.cell(row=row, column=3, value=job["licence"])
            ws.cell(row=row, column=4, value=job["request_no"])
            ws.cell(row=row, column=5, value=job["job_id"])
            ws.cell(row=row, column=6, value=job["jeweller"])
            ws.cell(row=row, column=7, value=pcs)
            ws.cell(row=row, column=8, value=weight)
            ws.cell(row=row, column=9, value=scrap_weight)
            ws.cell(row=row, column=10, value=current_weight)
            ws.cell(row=row, column=11, value=job["purity"])
            ws.cell(row=row, column=12, value=base_amount)
            ws.cell(row=row, column=13, value=gst_amount)
            ws.cell(row=row, column=14, value=cgst_amount)
            ws.cell(row=row, column=15, value=sgst_amount)
            ws.cell(row=row, column=16, value=total_amount)
            
            # Apply borders and alignment
            for col in range(1, 17):
                cell = ws.cell(row=row, column=col)
                cell.border = styles["border"]
                cell.alignment = styles["center_alignment"]
                if col in [8, 9, 10, 12, 13, 14, 15, 16]:  # Numeric columns
                    cell.number_format = '#,##0.00'
            
            row += 1
        
        # Grand totals row
        totals_row = row
        ws[f"A{totals_row}"] = "GRAND TOTALS"
        ws[f"A{totals_row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{totals_row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws[f"A{totals_row}"].border = styles["border"]
        
        # Round grand totals to 2 decimals
        grand_total_base_amount = round(grand_total_base_amount, 2)
        grand_total_gst = round(grand_total_gst, 2)
        grand_total_cgst = round(grand_total_cgst, 2)
        grand_total_sgst = round(grand_total_sgst, 2)
        
        # Calculate grand total amount
        grand_total_gst_calc = grand_total_base_amount * self.GST_RATE
        grand_total_amount = grand_total_base_amount + grand_total_gst_calc
        grand_total_amount = round(grand_total_amount, 2)
        
        ws.cell(row=totals_row, column=7, value=grand_total_pcs)
        ws.cell(row=totals_row, column=8, value=grand_total_weight)
        ws.cell(row=totals_row, column=9, value=grand_total_scrap)
        ws.cell(row=totals_row, column=10, value=grand_total_current)
        ws.cell(row=totals_row, column=12, value=grand_total_base_amount)
        ws.cell(row=totals_row, column=13, value=grand_total_gst)
        ws.cell(row=totals_row, column=14, value=grand_total_cgst)
        ws.cell(row=totals_row, column=15, value=grand_total_sgst)
        ws.cell(row=totals_row, column=16, value=grand_total_amount)
        
        for col in range(1, 17):
            cell = ws.cell(row=totals_row, column=col)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.border = styles["border"]
            cell.alignment = styles["center_alignment"]
            if col in [8, 9, 10, 12, 13, 14, 15, 16]:
                cell.number_format = '#,##0.00'
        
        # Add pricing rules
        row = totals_row + 3
        ws[f"A{row}"] = "PRICING RULES:"
        ws[f"A{row}"].font = Font(bold=True, size=10)
        
        row += 1
        ws[f"A{row}"] = "• PCS < 5: Fixed Rate = ₹200"
        
        row += 1
        ws[f"A{row}"] = "• PCS ≥ 5: Rate = ₹45 × PCS"
        
        row += 1
        ws[f"A{row}"] = "• GST 18% = CGST 9% + SGST 9%"
        
        # Generate filename
        timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"JobExport_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
